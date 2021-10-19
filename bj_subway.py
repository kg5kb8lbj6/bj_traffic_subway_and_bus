from pyquery import PyQuery
import requests
import datetime
from openpyxl import workbook
import re
# 进行数据的爬取
def sub_link():
    url = 'https://dt.8684.cn/bj'
    headers = {'User-Agent':
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
                }
    res = requests.get(url,headers = headers)
    #print('如果打印200则是请求成功:',res.status_code)
    start_time = datetime.datetime.now()
    print('开始爬取数据的时间为：',start_time)
    res.encoding = 'utf-8' # 对原网页的数据进行编码
    html_data = res.text
    subway_name = [] # 地铁
    subway_link = [] # 地铁的链接
    subway_name_list = [] # 最终的地铁路线名
    subway_link_list = []  #最终的地铁各个站台名
    doc =  PyQuery((''.join(html_data)))
    for item in doc.items('ul li a'):
        subway_name.append(item.text())
        subway_link.append(item.attr.href)
    # 对各个地铁进行处理
    subway_name = subway_name[:54]
    for i in range(1,len(subway_name)):
        if i % 2 != 0:
            subway_name_list.append(subway_name[i])
    # 对各个站台名进行处理
    subway_link = subway_link[:54]
    for i in range(1,len(subway_link)):
        if i % 2 != 0:
            subway_link_list.append(subway_link[i])
    # 发现地铁2号线和10号线与其他的结构是不一样的,所以把这两条线路单独放在一块,其他的线路放在一块
    subway_name_list_1 = [] # 除了2和10号线所有的地铁
    subway_link_1 = []
    subway_name_list_2_10 = []
    subway_link_2_10 = []
    for i in range(0, len(subway_link_list)):
        if i == 1 or i == 2 or i == 10 or i == 11:
            subway_name_list_2_10.append(subway_name_list[i])
            subway_link_2_10.append(subway_link_list[i])
        else:
            subway_name_list_1.append(subway_name_list[i])
            subway_link_1.append(subway_link_list[i])

    return subway_name_list_1,subway_link_1,subway_name_list_2_10,subway_link_2_10,start_time

def save(sub_name_1,sub_link_1,sub_name_2_10,subway_link_2_10):
    excel = workbook.Workbook() #创建一个excel文件
    sheet = excel.active
    # 先去求除2号线和10号线以后所有的地铁
    name_temp_1 = []
    name_1 = []
    station_name = [] # 总的站台

    star_time_1= [] # 首班车时间
    star_time_2 = []

    end_time_1 = [] # 尾班车时间
    end_time_2 = []

    station_name_2_10 = []

    station_temp_2_10 = [] # 2-10的站台名站台名
    str_time = []
    str_time_2_10 = [] # 2-10的时间

    final_station = []

    final_start_time_2_10 = [] # 最终的
    final_end_time_2_10 = [] 

    for k in range(0, len(sub_name_1)):  #len(sub_name_1)
        url = 'https://dt.8684.cn' + sub_link_1[k]
        station_1 = [] # 站台名
        time_total = []
        headers = {'User-Agent':
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
                }
        res = requests.get(url,headers = headers)
        #print('如果打印200则是请求成功:',res.status_code)
        res.encoding = 'utf-8' # 对原网页的数据进行编码
        html_data = res.text
        doc =  PyQuery((''.join(html_data)))

        for item in doc.items('div em'):
            name_temp_1.append(item.text())
        for item in doc.items('tr td a'):
            station_1.append(item.text())

        for item in doc.items('tr td'):
            time_total.append(item.text())
        str_time.append(time_total)
        station_name.append(station_1)
    for i in range(0, len(str_time)):
        star_temp_1= [] # 首班去的集合
        star_temp_2 = [] # 回来
        end_temp_1= []  # 尾班
        end_temp_2 = []
        for j in range(0, len(str_time[i])):
            if j % 5 == 1:
                star_temp_1.append(str_time[i][j])
            if j % 5 == 2:
                star_temp_2.append(str_time[i][j])
            if j % 5 == 3:
                end_temp_1.append(str_time[i][j])
            if j % 5 == 4:
                end_temp_2.append(str_time[i][j])
        star_time_1.append(star_temp_1) #  首班车去的时间
        star_time_2.append(star_temp_2)
        end_time_1.append(end_temp_1) # 尾班去的时间
        end_time_2.append(end_temp_2)

    for i in range(0,len(name_temp_1)):
        if i % 2 == 0:
            name_1.append(name_temp_1[i]) # name_1 地铁线路
    
    # 爬取剩余的2号线和10号线
    for r in range(0,len(sub_name_2_10)):
        url_1= 'https://dt.8684.cn' + subway_link_2_10[r]
        #print(url_1)
        
        time_total_2_10 = []
        headers = {'User-Agent':
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
                }
        res_1= requests.get(url_1,headers = headers)
        #print('如果打印200则是请求成功:',res.status_code)
        res_1.encoding = 'utf-8' # 对原网页的数据进行编码
        html_data_1 = res_1.text
        doc_1=  PyQuery((''.join(html_data_1)))
        for item in doc_1.items('div em'): # 站台名
            station_temp_2_10.append(item.text())
        for item in doc_1.items('tr td'):
            time_total_2_10.append(item.text())
        str_time_2_10.append(time_total_2_10)
    #print(str_time_2_10[0])
    #print(str_time_2_10[2])
    for i in range(0, len(str_time_2_10)): # 4
        station_2_10 = [] # 站台
        final_start_time_2 = [] # 开始时间
        final_end_time_2 = [] # 结束时间
        for j in range(0, len(str_time_2_10[i])):
            if j % 3 == 0:
                station_2_10.append(str_time_2_10[i][j])
            if j % 3 == 1:
                final_start_time_2.append(str_time_2_10[i][j])
            if j % 3 == 2:
                final_end_time_2.append(str_time_2_10[i][j])
        final_station.append(station_2_10) # 2,10号线的站名
        final_start_time_2_10.append(final_start_time_2) # 2,10 开始时间
        final_end_time_2_10.append(final_end_time_2) # 2,10 结束时间
    for i in range(1,len(station_temp_2_10)):
        if i % 3== 0:
            station_name_2_10.append(station_temp_2_10[i])
    
    for i in range(0,len(name_1)):
        sheet.cell(1, 5*(i+1)-4).value = name_1[i]
        sheet.cell(1, 5*(i+1)-3).value = '首班车'
        sheet.cell(1, 5*(i+1)-2).value = '首班车'
        sheet.cell(1, 5*(i+1)-1).value = '尾班车'
        sheet.cell(1, 5*(i+1)).value = '尾班车'
    #print(sheet.max_column)
    for i in range(0,len(station_name)): # len(station_name) = 23
        for j in range(0,len(station_name[i])):
            sheet.cell(j + 2,5*(i+1)-4).value = station_name[i][j] # 站台名

    for i in range(0,len(star_time_1)):
        for j in range(0,len(star_time_1[i])):
            sheet.cell(j+2,5*(i+1)-3).value = star_time_1[i][j]
            sheet.cell(j+2,5*(i+1)-2).value = star_time_2[i][j]
            sheet.cell(j+2,5*(i+1)-1).value = end_time_1[i][j]
            sheet.cell(j+2,5*(i+1)).value = end_time_2[i][j]


    '''对2号线和10号线的写入以及爬取可以进一步的优化'''

    
    # 前面一共115列
    # 将2号线和10号线写入文件中
    sheet.cell(1,116).value = sub_name_2_10[0]
    sheet.cell(1,117).value = '首班车'
    sheet.cell(1,118).value = '尾班车'
    sheet.cell(1,119).value = sub_name_2_10[1]
    sheet.cell(1,120).value = '首班车'
    sheet.cell(1,121).value = '尾班车'
    sheet.cell(1,122).value = sub_name_2_10[2]
    sheet.cell(1,123).value = '首班车'
    sheet.cell(1,124).value = '尾班车'
    sheet.cell(1,125).value = sub_name_2_10[3]
    sheet.cell(1,126).value = '首班车'
    sheet.cell(1,127).value = '尾班车'
    # final_station:2,10号线的站名  
    # final_start_time_2_10:2,10 开始时间 
    # final_end_time_2_10:2,10 结束时间
    for i in range(0,len(final_station)):
        for j in range(0,len(final_station[i])):
            sheet.cell(j+2,3*(i+1)+113).value = final_station[i][j]
    for i in range(0,len(final_start_time_2_10)):
        for j in range(0,len(final_start_time_2_10[i])):
            sheet.cell(j+2,3*(i+1)+114).value = final_start_time_2_10[i][j]
    for i in range(0,len(final_end_time_2_10)):
        for j in range(0,len(final_end_time_2_10[i])):
            sheet.cell(j+2,3*(i+1)+115).value = final_end_time_2_10[i][j]

    excel.save(r'D:/bj_subway.xlsx')
if __name__ == "__main__":
    print('开始爬取,################保持微笑,不要出错#############')
    sub_name_1,sub_link_1,sub_name_2_10,subway_link_2_10,start_time= sub_link()
    save(sub_name_1,sub_link_1,sub_name_2_10,subway_link_2_10)
    cur_time = datetime.datetime.now()
    print('爬取结束的时间为:',cur_time)
    cost_time = cur_time - start_time
    print('爬取共用的时间为:',cost_time)
    print('终于结束,可以大笑!！！！！！！！！！！！！！！！！！！')


