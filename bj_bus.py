from pyquery import PyQuery
import requests
import datetime
from openpyxl import workbook

# 通用的获取链接的页面
def common_link(url):
    url = url
    headers = {'User-Agent':
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
                }
    res = requests.get(url, headers = headers)
    res.encoding = 'utf-8'
    html_data = res.text
    doc =  PyQuery((''.join(html_data)))
    return doc
    
def get_url(): # 获取了13个大类的链接
    bus_url_temp = [] # 暂时存放
    bus_url = []
    doc = common_link('https://m.8684.cn/beijing_buslist')
    for item in doc.items('div a'):
        bus_url_temp.append(item.attr.href)
    bus_url_temp = bus_url_temp[2:15]
    for i in range(0, len(bus_url_temp)):
        bus_url.append('https://m.8684.cn'+bus_url_temp[i])
    print('第一部分当前时间:',datetime.datetime.now())
    return bus_url

# 获取每一辆公交的name和link
def every_bus_name_and_url():
    every_bus_name = []
    every_bus_link = []
    bus_url = get_url()
    for i in range(0, len(bus_url)): # len(bus_url)
        every_bus_name_temp = []
        every_bus_link_temp = []
        doc = common_link(bus_url[i])
        for item in doc.items('div.col-4.fold-inner a'):
            every_bus_name_temp.append(item.text())
            every_bus_link_temp.append('https://m.8684.cn'+item.attr.href)
        every_bus_name.append(every_bus_name_temp)
        every_bus_link.append(every_bus_link_temp)
    return every_bus_name, every_bus_link

# 获取每一辆公交车所经过的站牌
def stop_name():
    every_bus_name, every_bus_link = every_bus_name_and_url()
    station_name = []
    for i in range(0, len(every_bus_link)): # len(every_bus_link)
        for j in range(0, len(every_bus_link[i])):
            station_name_temp = []
            doc = common_link(every_bus_link[i][j])
            for item in doc.items('div.cell-group.show ol li a span.place'):
                station_name_temp.append(item.text())
            #print(station_name_temp)
            station_name.append(station_name_temp)
    return station_name,every_bus_name

# 保存信息到xlsx文件
def save_xlsx():
    excel = workbook.Workbook()
    sheet = excel.active
    station_name, every_bus_name = stop_name()
    k = 0
    for i in range(0, len(every_bus_name)): # bus name
        for j in range(0, len(every_bus_name[i])):
            sheet.cell(1+j+k, 1).value = every_bus_name[i][j]
        k += len(every_bus_name[i])
    # print(station_name)
    # print(len(station_name))
    # print(len(station_name[0]))
    # print(station_name[0][0])
    for i in range(0, len(station_name)):
        for j in range(0, len(station_name[i])):
            sheet.cell(1+i,2 + j).value = station_name[i][j]
    excel.save(r'D:/bj_bus.xlsx')

if __name__ == '__main__':
    start_time = datetime.datetime.now()
    print('开始爬取的时间为:',start_time)
    save_xlsx()
    end_time = datetime.datetime.now()
    print('爬取结束的时间为:',end_time)
    cost_time = end_time - start_time
    print('一共消耗了:',cost_time)
    
